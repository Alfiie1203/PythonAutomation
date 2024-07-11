import os # Importa el módulo os para trabajar con funciones del sistema operativo
from reportlab.lib.pagesizes import letter # Importa el tamaño de página letter de reportlab
from reportlab.lib import utils # Importa utilidades de reportlab
from reportlab.pdfgen import canvas # Importa canvas para generar PDFs
import pandas as pd # Importa pandas para manipulación de datos
import openpyxl # Importa openpyxl para trabajar con archivos Excel
import threading # Importa threading para trabajar con hilos
import time # Importa time para funciones de tiempo
from openpyxl.styles import PatternFill # Importa PatternFill para estilos en celdas de Excel
import win32com.client as win32 # Importa win32com para automatización de aplicaciones COM (como Excel)
from Bot_i129s import generate_ExcelKey # Importa la función generate_ExcelKey del módulo Bot_i129s
from ToolBook import utils
# Variables globales para controlar el estado del bot
bot_running = False
bot_thread = None

# Rutas base para archivos y carpetas
ruta_base_user = 'G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/USERS/'
ruta_base = 'G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/BOT - DO NOT TOUCH/'
templates = ruta_base+'templates/'

output_folder = ruta_base_user+'pdfs_generados'

# Crear el directorio de salida si no existe
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

def generate_pdfs_from_excel(excel_file, image_paths, output_folder):
    """
    Genera PDFs a partir de datos en un archivo Excel.
    """
    # Leer el archivo Excel
    xls = pd.ExcelFile(excel_file)
    data_temp = pd.read_excel(xls, 'Temp')
    data_log = pd.read_excel(xls, 'Log')
    
    # Reemplazar NaN por cadenas vacías
    data_temp = data_temp.fillna('')

    # Renombrar columnas problemáticas temporalmente
    data_temp = data_temp.rename(columns={
        'Vía pública': 'Via publica',
    })
    
    # Definir los campos obligatorios
    required_fields = [
        "Name of the Petitioning Organization", "In Care Of Name (if any) last", 
        "In Care Of Name (if any) first", "U.S. Street address", "City_Petitioner", 
        "State_Petitioner", "Zip Code_Petitioner", 
        "Is this mailing address the same as the physical location of the sponsoring company or organization?", 
        "Daytime Telephone Number", "Email Address (if any)", "Website Address (if any)", 
        "Does the petitioner employ 50 or more individuals in the United States?", 
        "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?", 
        "The beneficiary will work as a:", "startdate_proposed_employment", 
        "enddate_proposed_employment", "Was the beneficiary of this petition in the United States during the last seven years?", 
        "Family Name (Last Name)", "Given Name (First Name)", "Street Number and Name or Po Box", 
        "City_Beneficiary", "Province_Beneficiary", "PostalCode_Beneficiary", "Country_Beneficiary", 
        "Date of Birth", "Gender", "City of birth", "State of birth", "Country of birth", 
        "Country of Citizenship or Nationality", "Number for the Blanket", 
        "Beneficiary's Wages Per Year", "Beneficiary's Hours Per Week", "Job Title", 
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer", 
        "foreign Employer Name", "Street Address Mailing Address", "City Mailing Address", 
        "Province Mailing Address", "Postal Code Mailing Address", "Country Mailing Address", 
        "Job Title_Act", "Start Date", "Wages Earned Per Year", "Hours Worked Per Week", 
        "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:", 
        "Petitioner's or Authorized Signatory's Title", "Preparer's Family Name (Last Name)", 
        "Preparer's Given Name (First Name)", "Preparer's Business or Organization Name", 
        "Preparer's Daytime Telephone Number", "Preparer's Email Address (if any)"
    ]

    for index, row in data_temp.iterrows():
        
        # Verificar campos obligatorios
        missing_fields = [field for field in required_fields if pd.isna(row[field]) or row[field] == '']
        
        if missing_fields:
            #print(f"\n The line that corresponds to the company {str(row['Name of the Petitioning Organization'])} and the user {str(row['Middle Name'])} {str(row['Family Name (Last Name)'])}, {str(row['Given Name (First Name)'])} which was located in the line: {index + 1} was omitted due to missing fields: \n {missing_fields} \n \n")
            continue
        
        # Nombre del archivo PDF de salida
        output_pdf = f'{output_folder}/Modelo_i-129s_{str(row['Name of the Petitioning Organization'])}_{str(row['Middle Name'])} {str(row['Family Name (Last Name)'])}, {str(row['Given Name (First Name)'])}.pdf'

        # Crear un lienzo PDF
        c = canvas.Canvas(output_pdf, pagesize=letter)
        
        # Crear el JSON de los datos de texto
        texto_data_hoja1 = [
            #Trabajador ----->
            {"text": str(row['Name of the Petitioning Organization']), "x": 65, "y": 402},
            {"text": str(row['In Care Of Name (if any) first'])+" "+str(row['In Care Of Name (if any) last']), "x": 65, "y": 336},
            {"text": str(row['U.S. Street address']), "x": 129, "y": 311},
            #{"text": str(row['h1_2c']), "x": 55, "y": 657}, hace falta confirmar la manera
            {"text": str(row['City_Petitioner']), "x": 129, "y": 264},
            {"text": str(row['State_Petitioner']), "x": 87, "y": 239},
            {"text": str(row['Zip Code_Petitioner']), "x": 195, "y": 240},
            #{"text": str(row['h1_3']), "x": 218, "y": 193}, Se hace como condicion yes/no
            {"text": str(row['Daytime Telephone Number']), "x": 347, "y": 287},
            #{"text": str(row['h1_6']), "x": 347, "y": 252},
            {"text": str(row['Email Address (if any)']), "x": 347, "y": 215},
            {"text": str(row['Website Address (if any)']), "x": 347, "y": 180},
            #{"text": str(row['h1_9']), "x": 55, "y": 657}, Se hace como condicion yes/no
        ]
        
        if(str(row['Is this mailing address the same as the physical location of the sponsoring company or organization?']) == "YES"):  
            texto_data_hoja1.append({"text": "X", "x": 217, "y": 191})
        elif(str(row['Is this mailing address the same as the physical location of the sponsoring company or organization?']) == "NO"):
            texto_data_hoja1.append({"text": "X", "x": 259, "y": 191})
        
        if(str(row['Does the petitioner employ 50 or more individuals in the United States?']) == "YES"):  
            texto_data_hoja1.append({"text": "X", "x": 499, "y": 113})
        elif(str(row['Does the petitioner employ 50 or more individuals in the United States?']) == "NO"):
            texto_data_hoja1.append({"text": "X", "x": 541, "y": 113})
        
        texto_data_hoja2 = [
            #Trabajador ----->
            #{"text": str(row['h2_10']), "x": 91, "y": 577},
            #{"text": str(row['h2_1a']), "x": 65, "y": 577},
            #{"text": str(row['h2_1b']), "x": 91, "y": 577},
            {"text": utils.convert_date_format(str(row['startdate_proposed_employment'])), "x": 212, "y": 414},
            {"text": utils.convert_date_format(str(row['enddate_proposed_employment'])), "x": 212, "y": 389},
            #{"text": str(row['h2_3']), "x": 91, "y": 577},
            {"text": str(row['Family Name (Last Name)']), "x": 406, "y": 432},
            {"text": str(row['Given Name (First Name)']), "x": 406, "y": 408},
            {"text": str(row['Middle Name']), "x": 406, "y": 383},
        ]
        
        if(str(row["Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?"]) == "YES"):
            texto_data_hoja2.append({"text": "X", "x": 217, "y": 665})
        elif(str(row["Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?"]) == "NO"):
            texto_data_hoja2.append({"text": "X", "x": 259, "y": 665})
            
        if(str(row['The beneficiary will work as a:']) == "Manager or Executive"):
            texto_data_hoja2.append({"text": "X", "x": 62, "y": 502})
        elif(str(row['The beneficiary will work as a:']) == "Specialized Knowledge Professional"):
            texto_data_hoja2.append({"text": "X", "x": 62, "y": 485})
            
        #if(str(row['H1_parte2_N1b']) == "YES"):  
        #    texto_data_hoja2.append({"text": "X", "x": 62, "y": 485})
        #elif(str(row['H1_parte2_N1b']) == "NO"):
        #    texto_data_hoja2.append({"text": "", "x": 62, "y": 485})
        
        if(str(row['Was the beneficiary of this petition in the United States during the last seven years?']) == "YES"):  
            texto_data_hoja2.append({"text": "X", "x": 217, "y": 323})
        elif(str(row['Was the beneficiary of this petition in the United States during the last seven years?']) == "NO"):
            texto_data_hoja2.append({"text": "X", "x": 259, "y": 323})
            
        
        texto_data_hoja3 = [
            #Trabajador ----->
            {"text": str(row['Given Name (First Name)'])+" "+ str(row['Middle Name'])+" "+ str(row['Family Name (Last Name)']), "x": 62, "y": 647},
            {"text": str(row['Street Number and Name or Po Box']), "x": 62, "y": 612},
            {"text": str(row['City_Beneficiary']), "x": 127, "y": 564},	
            {"text": str(row['Province_Beneficiary']), "x": 127, "y": 539},	
            {"text": str(row['PostalCode_Beneficiary']), "x": 127, "y": 516},	
            {"text": str(row['Country_Beneficiary']), "x": 61, "y": 480},
            	
            {"text": str(utils.convert_date_format(row['Date of Birth'])), "x": 494, "y": 701},	
            #{"text": str(row['h3_10']), "x": 91, "y": 577},	 
            {"text": str(row['City of birth']), "x": 344, "y": 640},	
            {"text": str(row['State of birth']), "x": 344, "y": 605},	
            {"text": str(row['Country of birth']), "x": 344, "y": 570},	
            {"text": str(row['Country of Citizenship or Nationality']), "x": 344, "y": 534},
            
            {"text": utils.addSpaces(str(row['Number for the Blanket'])), "x": 391, "y": 432, "is_add_spaces": True},
            {"text": str(row['U.S. Street address']), "x": 411, "y": 348},
            {"text": str(row['City_Petitioner']), "x": 411, "y": 300},
            {"text": str(row['State_Petitioner']), "x": 370, "y": 275},
            {"text": str(row['Zip Code_Petitioner']), "x": 477, "y": 275},
            
            {"text": str(row["Beneficiary's Wages Per Year"]), "x": 479, "y": 154},	
            {"text": str(row["Beneficiary's Hours Per Week"]), "x": 479, "y": 132},	

        ]
        
        if(str(row['Gender']) == "MALE"):  
            texto_data_hoja3.append({"text": "X", "x": 387, "y": 678})
        elif(str(row['Gender']) == "FEMALE"):
            texto_data_hoja3.append({"text": "X", "x": 442, "y": 678})
        
        texto_data_hoja4 = [
            {"text": str(row['Job Title']), "x": 61, "y": 575},
            #{"text": str(row['h4_1a_b_c']), "x": 411, "y": 348},
            {"text": str(row['foreign Employer Name']), "x": 345, "y": 443},
            {"text": str(row['Street Address Mailing Address']), "x": 410, "y": 390},
            {"text": str(row['City Mailing Address']), "x": 410, "y": 342},
            {"text": str(row['Province Mailing Address']), "x": 410, "y": 317},
            {"text": str(row['Postal Code Mailing Address']), "x": 410, "y": 293},
            {"text": str(row['Country Mailing Address']), "x": 344, "y": 257},

        ]
        texto_data_hoja4.append({"text": "Please refer to support letter.", "x": 61, "y": 539})
                
        if(str(row['Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer']) == "Manager"):  
            texto_data_hoja4.append({"text": "X", "x": 345, "y": 576})
        elif(str(row['Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer']) == "Executive"):
            texto_data_hoja4.append({"text": "X", "x": 345, "y": 557})
        elif(str(row['Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer']) == "Specialized Knowledge Professional"):
            texto_data_hoja4.append({"text": "X", "x": 345, "y": 539})
            
            
            
        texto_data_hoja5 = [
            {"text": str(row['Job Title_Act']), "x": 61, "y": 552},
            {"text": utils.convert_date_format(str(row['Start Date'])), "x": 211, "y": 527},
            {"text": str(row['Wages Earned Per Year']), "x": 193, "y": 408},
            {"text": str(row['Hours Worked Per Week']), "x": 193, "y": 383}, #hoja 1
            #Condicionar
            #{"text": str(row['With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:']), "x": 61, "y": 575}, Condicional 1/2 = x
            {"text": str(row['In Care Of Name (if any) last']), "x": 345, "y": 293},
            {"text": str(row['In Care Of Name (if any) first']), "x": 345, "y": 245},
            {"text": str(row["Petitioner's or Authorized Signatory's Title"]), "x": 345, "y": 209},
            {"text": str(row['Daytime Telephone Number']), "x": 345, "y": 161},
            {"text": str(row['Email Address (if any)']), "x": 345, "y": 65},

        ]
        texto_data_hoja5.append({"text": "Please refer to support letter.", "x": 61, "y": 464})
        
        texto_data_hoja6 = []
        texto_data_hoja6.extend([
            {"text": str(row["Preparer's Family Name (Last Name)"]), "x": 346, "y": 408},
            {"text": str(row["Preparer's Given Name (First Name)"]), "x": 346, "y": 372},
            {"text": str(row["Preparer's Business or Organization Name"]), "x": 346, "y": 336},
            {"text": str(row["Preparer's Daytime Telephone Number"]), "x": 346, "y": 269},
            {"text": str(row["Preparer's Email Address (if any)"]), "x": 346, "y": 197}
        ])
        
        
        texto_data_hoja7 = []
        
        texto_data_hoja8 = []
        
        
             
        utils.add_text_to_image(c, image_paths[0], texto_data_hoja1)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[1], texto_data_hoja2)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[2], texto_data_hoja3)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[3], texto_data_hoja4)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[4], texto_data_hoja5)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[5], texto_data_hoja6)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[6], texto_data_hoja7)
        c.showPage()  # Añadir nueva página para la siguiente imagen
        
        utils.add_text_to_image(c, image_paths[7], texto_data_hoja8)
        c.showPage()  # Añadir nueva página para la siguiente imagen

        # Guardar el PDF
        c.save()

        #Mover la fila de Temp a Log
        data_log = pd.concat([data_log, pd.DataFrame([row])])
        data_temp = data_temp.drop(index)

    # Guardar los cambios en el archivo Excel
    with pd.ExcelWriter(excel_file, mode='a', if_sheet_exists='replace') as writer:
        data_temp.to_excel(writer, sheet_name='Temp', index=False)
        data_log.to_excel(writer, sheet_name='Log', index=False)
        
    # Colorear las celdas vacías en la hoja "Temp"
    workbook = openpyxl.load_workbook(excel_file)
    utils.color_cells(workbook, 'Temp')
    workbook.save(excel_file)
        
def process_files():
    global bot_running # Indica que usaremos la variable global bot_running
    while bot_running: # Bucle que se ejecuta mientras el bot esté activo
        # Ruta del archivo Excel de entrada
        excel_file = ruta_base+'INPUT USERS DATA FORM I-129S.xlsx'
        # Rutas de las imágenes de las plantillas de las páginas del PDF
        image_paths = [templates+'page_1.jpg', 
                       templates+'page_2.jpg',
                       templates+'page_3.jpg',
                       templates+'page_4.jpg',
                       templates+'page_5.jpg',
                       templates+'page_6.jpg',
                       templates+'page_7.jpg',
                       templates+'page_8.jpg']
        
        # Genera la clave de Excel necesaria para el procesamiento
        generate_ExcelKey.generateExcelKey()
        # Genera los PDFs a partir del archivo Excel y las imágenes de las plantillas
        generate_pdfs_from_excel(excel_file, image_paths, output_folder)
        # Espera 5 minutos (300 segundos) antes de volver a comprobar
        time.sleep(300)

def start_bot():
    global bot_running, bot_thread  # Indica que usaremos las variables globales bot_running y bot_thread
    if not bot_running:  # Si el bot no está corriendo actualmente
        bot_running = True  # Marca el bot como corriendo
        # Crea y empieza un nuevo hilo que ejecuta la función process_files
        bot_thread = threading.Thread(target=process_files)
        bot_thread.start()

def stop_bot():
    global bot_running, bot_thread  # Indica que usaremos las variables globales bot_running y bot_thread
    if bot_running:  # Si el bot está corriendo actualmente
        bot_running = False  # Marca el bot como no corriendo
        if bot_thread:  # Si hay un hilo de bot activo
            bot_thread.join()  # Espera a que el hilo termine
            bot_thread = None  # Resetea la variable del hilo

#excel_file = ruta_base+'INPUT USERS DATA FORM I-129S.xlsx'
#image_paths = ['G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_1.jpg', 
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_2.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_3.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_4.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_5.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_6.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_7.jpg',
#               'G:/Shared drives/ES VIALTO GMS - RPA/TAX/COMPLIANCE/i_129s/templates/page_8.jpg']
#
#generate_pdfs_from_excel(excel_file, image_paths, output_folder)


