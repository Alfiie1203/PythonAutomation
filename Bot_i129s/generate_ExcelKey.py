import pandas as pd
import os
from openpyxl import load_workbook

# Ruta de la carpeta con los archivos de Excel
ruta_archivos = "G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/BOT - DO NOT TOUCH/OUT FASE 1 (G sheet to excel)/"
# Nombre del archivo a actualizar
archivo_actualizar = "G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/BOT - DO NOT TOUCH/INPUT USERS DATA FORM I-129S.xlsx"


def generateExcelKey():
    # Crear un DataFrame vacío para almacenar los datos
    df_total = pd.DataFrame()

    # Recorrer todos los archivos en la carpeta
    for archivo in os.listdir(ruta_archivos):
        if archivo.endswith('.xlsx'):
            archivo_path = os.path.join(ruta_archivos, archivo)
            # Leer la hoja "Temp" de cada archivo
            try:
                df_temp = pd.read_excel(archivo_path, sheet_name='Temp')
                df_total = pd.concat([df_total, df_temp], ignore_index=True)
                # Eliminar el archivo después de procesarlo
                os.remove(archivo_path)
                print(f"Archivo procesado y eliminado: {archivo}")
            except Exception as e:
                print(f"No se pudo leer la hoja 'Temp' del archivo {archivo}. Error: {e}")

    # Leer el archivo de destino
    with pd.ExcelFile(archivo_actualizar) as xls:
        # Leer todas las hojas del archivo existente
        hojas = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}

    # Crear un DataFrame con la hoja "Temp" existente si es que ya existe
    if 'Temp' in hojas:
        df_existente = hojas['Temp']
        # Concatenar los datos existentes con los nuevos
        df_actualizado = pd.concat([df_existente, df_total], ignore_index=True)
    else:
        df_actualizado = df_total

    # Abrir el archivo con openpyxl para la actualización
    libro = load_workbook(archivo_actualizar)

    # Eliminar la hoja 'Temp' si ya existe
    if 'Temp' in libro.sheetnames:
        std = libro['Temp']
        libro.remove(std)
        libro.save(archivo_actualizar)

    # Guardar los datos actualizados en la hoja "Temp"
    with pd.ExcelWriter(archivo_actualizar, engine='openpyxl', mode='a') as writer:
        df_actualizado.to_excel(writer, sheet_name='Temp', index=False)

    
    