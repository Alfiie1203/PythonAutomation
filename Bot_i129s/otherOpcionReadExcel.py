import pandas as pd
import os
import difflib
import shutil
import threading
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

bot_running = False
bot_thread = None
ruta_base = 'G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/'
ruta_user = ruta_base+'USERS/'
ruta_admin = ruta_base+'BOT - DO NOT TOUCH/'
ruta_out = ruta_admin+'IN FASE 1 (d60 to excels)/'
templates = ruta_base+'BOT - DO NOT TOUCH/templates/'

def diccionario():
    df = pd.read_excel(ruta_admin+"diccionario.xlsx")
    return df

def preprocess_string(s):
    return s.lower().strip()

def fill_data(df):
    selected_columns = [
        "Name of the Petitioning Organization",
        "In Care Of Name (if any) last",
        "In Care Of Name (if any) first",
        "U.S. Street address",
        "City_Petitioner",
        "State_Petitioner",
        "Zip Code_Petitioner",
        "Is this mailing address the same as the physical location of the sponsoring company or organization?",
        "Daytime Telephone Number",
        "Email Address (if any)",
        "Website Address (if any)",
        "Does the petitioner employ 50 or more individuals in the United States?",
        "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?",
        "The beneficiary will work as a:",
        "startdate_proposed_employment",
        "enddate_proposed_employment",
        "Was the beneficiary of this petition in the United States during the last seven years?",
        "Family Name (Last Name)",
        "Given Name (First Name)",
        "Middle Name",
        "Street Number and Name or Po Box",
        "City_Beneficiary",
        "Province_Beneficiary",
        "PostalCode_Beneficiary",
        "Country_Beneficiary",
        "Date of Birth",
        "Gender",
        "City of birth",
        "State of birth",
        "Country of birth",
        "Country of Citizenship or Nationality",
        "Number for the Blanket",
        "Beneficiary's Wages Per Year",
        "Beneficiary's Hours Per Week",
        "Job Title",
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer",
        "foreign Employer Name",
        "Street Address Mailing Address",
        "City Mailing Address",
        "Province Mailing Address",
        "Postal Code Mailing Address",
        "Country Mailing Address",
        "Job Title_Act",
        "Start Date",
        "Wages Earned Per Year",
        "Hours Worked Per Week",
        "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that:",
        "Petitioner's or Authorized Signatory's Title",
    ]

    new_df = df[selected_columns]
    dic = diccionario()
    dic['preprocessed_organization'] = dic['Name of the Petitioning Organization'].apply(preprocess_string)

    for index, row in df.iterrows():
        preprocessed_value = preprocess_string(row['foreign Employer Name'])
        closest_match = difflib.get_close_matches(preprocessed_value, dic['preprocessed_organization'], n=1, cutoff=0.3)
        
        if closest_match:
            matched_value = closest_match[0]
            matched_row = dic[dic['preprocessed_organization'] == matched_value].iloc[0]
            # Asignar valores específicos a las columnas seleccionadas
            for col in selected_columns:
                if col in matched_row:
                    new_df.at[index, col] = matched_row[col]
        else:
            # Dejar la celda vacía si no se encuentra ninguna coincidencia
            for col in selected_columns:
                new_df.at[index, col] = ''

    return new_df

def color_cells(workbook, sheet_name, empty_columns_yellow, empty_columns_red):
    ws = workbook[sheet_name]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        column_letter = col[0].column_letter
        header = col[0].value

        if header in empty_columns_yellow:
            fill = yellow_fill
        else:
            fill = red_fill

        for cell in col:
            if cell.row > 1 and not cell.value:  # Verifica si la celda está vacía
                cell.fill = fill
                
def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def apply_header_style(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def apply_data_style(ws):
    data_font = Font(name='Calibri', size=11)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            
def add_dropdown(ws, col_letter, options):
    # Creamos el objeto DataValidation con tipo "list" y opciones especificadas
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', showDropDown=True)
    
    # Configuramos el mensaje de error
    dv.errorTitle = "Entrada inválida"
    dv.error = f"Debes seleccionar una opción de la siguiente lista: {'\n '.join(options)}"
    dv.showErrorMessage = True
    
    # Agregamos la validación de datos a la hoja de trabajo (worksheet)
    ws.add_data_validation(dv)
    
    # Aplicamos la validación de datos a todas las celdas en la columna especificada
    # desde la fila 2 hasta la última fila en la hoja de trabajo
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    
def generateExcel():
    carpeta = ruta_user+'/Casos/'
    datos_a_obtener = {
        "Name of the Petitioning Organization": "Name of the Petitioning Organization",
        "In Care Of Name (if any) last": "In Care Of Name (if any) last",
        "In Care Of Name (if any) first": "In Care Of Name (if any) first",
        "U.S. Street address": "U.S. Street address",
        "City_Petitioner": "City_Petitioner",
        "State_Petitioner": "State_Petitioner",
        "Zip Code_Petitioner": "Zip Code_Petitioner",
        "Is this mailing address the same as the physical location of the sponsoring company or organization?": "Is this mailing address the same as the physical location of the sponsoring company or organization?",
        "Daytime Telephone Number": "Daytime Telephone Number",
        "Email Address (if any)": "Email Address (if any)",
        "Website Address (if any)": "Website Address (if any)",
        "Does the petitioner employ 50 or more individuals in the United States?": "Does the petitioner employ 50 or more individuals in the United States?",
        "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?": "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?",
        "The beneficiary will work as a:": "The beneficiary will work as a:",
        "startdate_proposed_employment": "startdate_proposed_employment",
        "enddate_proposed_employment": "enddate_proposed_employment",
        "Was the beneficiary of this petition in the United States during the last seven years?": "Was the beneficiary of this petition in the United States during the last seven years?",
        "Family Name (Last Name)": "Family Name (Last Name)",
        "Given Name (First Name)": "Given Name (First Name)",
        "Middle Name": "Middle Name",
        "Street Number and Name or Po Box": "Street Number and Name or Po Box",
        "City_Beneficiary": "City_Beneficiary",
        "Province_Beneficiary": "Province_Beneficiary",
        "PostalCode_Beneficiary": "PostalCode_Beneficiary",
        "Country_Beneficiary": "Country_Beneficiary",
        "Date of Birth": "Date of Birth",
        "Gender": "Gender",
        "City of birth": "City of birth",
        "State of birth": "State of birth",
        "Country of birth": "Country of birth",
        "Country of Citizenship or Nationality": "Country of Citizenship or Nationality",
        "Number for the Blanket": "Number for the Blanket",
        "Beneficiary's Wages Per Year": "Beneficiary's Wages Per Year",
        "Beneficiary's Hours Per Week": "Beneficiary's Hours Per Week",
        "Job Title": "Job Title",
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer": "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer",
        "foreign Employer Name": "foreign Employer Name",
        "Street Address Mailing Address": "Street Address Mailing Address",
        "City Mailing Address": "City Mailing Address",
        "Province Mailing Address": "Province Mailing Address",
        "Postal Code Mailing Address": "Postal Code Mailing Address",
        "Country Mailing Address": "Country Mailing Address",
        "Job Title_Act": "Job Title_Act",
        "Start Date": "Start Date",
        "Wages Earned Per Year": "Wages Earned Per Year",
        "Hours Worked Per Week": "Hours Worked Per Week",
        "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that:": "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that:",
        "Petitioner's or Authorized Signatory's Title": "Petitioner's or Authorized Signatory's Title",
    }

    excels = [f for f in os.listdir(carpeta) if f.endswith('.xlsx') and not f.startswith('~$')]
    if not excels:
        print("No hay archivos Excel en la carpeta.")
        return
    
    for archivo in excels:
        filepath = os.path.join(carpeta, archivo)
        df = pd.read_excel(filepath)
        
        if 'Name of the Petitioning Organization' not in df.columns:
            print(f"No se encontró la columna 'Name of the Petitioning Organization' en el archivo {archivo}.")
            continue

        df_filled = fill_data(df)
        
        output_file = os.path.join(ruta_out, f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{archivo}")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_filled.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            ws = writer.sheets['Sheet1']
            
            color_cells(workbook, 'Sheet1', empty_columns_yellow=datos_a_obtener.keys(), empty_columns_red=[])
            adjust_column_width(ws)
            apply_header_style(ws)
            apply_data_style(ws)
            
            dropdown_options = {
                'H': ['Option 1', 'Option 2'],
                'J': ['Yes', 'No']
            }
            for col_letter, options in dropdown_options.items():
                add_dropdown(ws, col_letter, options)
                
    print(f"Los archivos se han procesado y guardado en la carpeta {ruta_out}")

# Llamada de prueba para la función
generateExcel()
