import pandas as pd
import os
import difflib
import shutil
import threading
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


bot_running = False
bot_thread = None
ruta_base = 'G:/Shared drives/ES VIALTO GMS - RPA/INMI & SS/FORM i129S/'
ruta_user = ruta_base+'USERS/'
ruta_admin = ruta_base+'BOT - DO NOT TOUCH/'
ruta_out = ruta_admin+'IN FASE 1 (d60 to excels)/'
templates = ruta_base+'BOT - DO NOT TOUCH/templates/'

def diccionario():
    df = pd.read_excel(ruta_admin+"Companies.xlsx")
    return df

def preprocess_string(s):
    if isinstance(s, str):
        return s.lower().strip()
    else:
        return ''

def fill_data(df):
    selected_columns = [
        "Name of the Petitioning Organization",
        "In Care Of Name (if any) last",
        "In Care Of Name (if any) first",
        "U.S. Street address",
        "City_Petitioner",
        "State_Petitioner",
        "Zip Code_Petitioner",
        "Is this mailing address the same as the physical location of the sponsoring company or organization?",
        "Daytime Telephone Number",
        "Email Address (if any)",
        "Website Address (if any)",
        "Does the petitioner employ 50 or more individuals in the United States?",
        "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?",
        "The beneficiary will work as a:",
        "startdate_proposed_employment",
        "enddate_proposed_employment",
        "Was the beneficiary of this petition in the United States during the last seven years?",
        "Family Name (Last Name)",
        "Given Name (First Name)",
        "Middle Name",
        "Street Number and Name or Po Box",
        "City_Beneficiary",
        "Province_Beneficiary",
        "PostalCode_Beneficiary",
        "Country_Beneficiary",
        "Date of Birth",
        "Gender",
        "City of birth",
        "State of birth",
        "Country of birth",
        "Country of Citizenship or Nationality",
        "Number for the Blanket",
        "Beneficiary's Wages Per Year",
        "Beneficiary's Hours Per Week",
        "Job Title",
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer",
        "foreign Employer Name",
        "Street Address Mailing Address",
        "City Mailing Address",
        "Province Mailing Address",
        "Postal Code Mailing Address",
        "Country Mailing Address",
        "Job Title_Act",
        "Start Date",
        "Wages Earned Per Year",
        "Hours Worked Per Week",
        "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:",
        "Petitioner's or Authorized Signatory's Title",
    ]

    new_df = df[selected_columns]
    dic = diccionario()
    dic['preprocessed_organization'] = dic['Name of the Petitioning Organization'].apply(preprocess_string)

    for index, row in df.iterrows():
        preprocessed_value = preprocess_string(row['foreign Employer Name'])
        closest_match = difflib.get_close_matches(preprocessed_value, dic['preprocessed_organization'], n=1, cutoff=0.3)
        
        if closest_match:
            matched_value = closest_match[0]
            matched_row = dic[dic['preprocessed_organization'] == matched_value].iloc[0]
            # Asignar valores específicos a las columnas seleccionadas
            new_df.at[index, 'Name of the Petitioning Organization'] = matched_row['Name of the Petitioning Organization']
            new_df.at[index, 'In Care Of Name (if any) first'] = matched_row['In Care Of Name (if any) first']
            new_df.at[index, 'In Care Of Name (if any) last'] = matched_row['In Care Of Name (if any) last']
            new_df.at[index, 'U.S. Street address'] = matched_row['U.S. Street address']
            new_df.at[index, 'City_Petitioner'] = matched_row['City_Petitioner']
            new_df.at[index, 'State_Petitioner'] = matched_row['State_Petitioner']
            new_df.at[index, 'Zip Code_Petitioner'] = matched_row['Zip Code_Petitioner']
            new_df.at[index, 'Is this mailing address the same as the physical location of the sponsoring company or organization?'] = matched_row['Is this mailing address the same as the physical location of the sponsoring company or organization?']
            new_df.at[index, 'Daytime Telephone Number'] = matched_row['Daytime Telephone Number']
            new_df.at[index, 'Email Address (if any)'] = matched_row['Email Address (if any)']
            new_df.at[index, 'Website Address (if any)'] = matched_row['Website Address (if any)']
            new_df.at[index, 'Does the petitioner employ 50 or more individuals in the United States?'] = matched_row['Does the petitioner employ 50 or more individuals in the United States?']
            new_df.at[index, "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?"] = matched_row["Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?"]
            new_df.at[index, "Beneficiary's Hours Per Week"] = matched_row["Beneficiary's Hours Per Week"]
            new_df.at[index, 'Number for the Blanket'] = matched_row['Number for the Blanket']
            new_df.at[index, "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:"] = matched_row["With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:"]
            new_df.at[index, "Petitioner's or Authorized Signatory's Title"] = matched_row["Petitioner's or Authorized Signatory's Title"]           
            
            new_df.at[index, "Preparer's Family Name (Last Name)"] = matched_row["Preparer's Family Name (Last Name)"]
            new_df.at[index, "Preparer's Given Name (First Name)"] = matched_row["Preparer's Given Name (First Name)"]
            new_df.at[index, "Preparer's Business or Organization Name"] = matched_row["Preparer's Business or Organization Name"]
            new_df.at[index, "Preparer's Daytime Telephone Number"] = matched_row["Preparer's Daytime Telephone Number"]
            new_df.at[index, "Preparer's Email Address (if any)"] = matched_row["Preparer's Email Address (if any)"]
        else:
            print(f"No match found in the folder")

    return new_df

def color_cells(workbook, sheet_name, empty_columns_yellow, empty_columns_red):
    ws = workbook[sheet_name]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        column_letter = col[0].column_letter
        header = col[0].value

        if header in empty_columns_yellow:
            fill = yellow_fill
        else:
            fill = red_fill

        for cell in col:
            if cell.row > 1 and not cell.value:  # Verifica si la celda está vacía
                cell.fill = fill
                
def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def apply_header_style(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def apply_data_style(ws):
    data_font = Font(name='Calibri', size=11)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            
def add_dropdown(ws, col_letter, options):
    # Creamos el objeto DataValidation con tipo "list" y opciones especificadas
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', showDropDown=True)
    
    # Configuramos el mensaje de error
    dv.errorTitle = "Entrada inválida"
    dv.error = f"Debes seleccionar una opción de la siguiente lista: {'\n '.join(options)}"
    dv.showErrorMessage = True
    
    # Agregamos la validación de datos a la hoja de trabajo (worksheet)
    ws.add_data_validation(dv)
    
    # Aplicamos la validación de datos a todas las celdas en la columna especificada
    # desde la fila 2 hasta la última fila en la hoja de trabajo
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

def generateExcel():
    carpeta = ruta_user+'Casos/'
    datos_a_obtener = {
        "Name of the Petitioning Organization": 1,
        "In Care Of Name (if any) last": 1,
        "In Care Of Name (if any) first": 1,
        "U.S. Street address": 1,
        "City_Petitioner": 1,
        "State_Petitioner": 1,
        "Zip Code_Petitioner": 1,
        "Is this mailing address the same as the physical location of the sponsoring company or organization?": 1,
        "Daytime Telephone Number": 1,
        "Email Address (if any)": 1,
        "Website Address (if any)": 1,
        "Does the petitioner employ 50 or more individuals in the United States?": 1,
        "Are more than 50 percent of the petitioner's employees in H-1B, L-1A, or L-1B nonimmigrant status?": 1,
        "The beneficiary will work as a:": 1,
        "startdate_proposed_employment": 1, #
        "enddate_proposed_employment": 1, #
        "Was the beneficiary of this petition in the United States during the last seven years?": 1, #
        "Family Name (Last Name)": 3,
        "Given Name (First Name)": 4,
        "Middle Name": 4,
        "Street Number and Name or Po Box": 85,
        "City_Beneficiary": 86,
        "Province_Beneficiary": 87,
        "PostalCode_Beneficiary": 88,
        "Country_Beneficiary": 89,
        "Date of Birth": 10,
        "Gender": 7,
        "City of birth": 11,
        "State of birth": 12,
        "Country of birth": 13,
        "Country of Citizenship or Nationality": 1, #
        "Number for the Blanket": 1, #Dic
        "Beneficiary's Wages Per Year": 1, #
        "Beneficiary's Hours Per Week": 1, #Dic
        "Job Title": 1, #
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer": 1, #
        "foreign Employer Name": 176,
        "Street Address Mailing Address": 178,
        "City Mailing Address": 179,
        "Province Mailing Address": 180,
        "Postal Code Mailing Address": 181,
        "Country Mailing Address": 183,
        "Job Title_Act": 175,
        "Start Date": 184,
        "Wages Earned Per Year": 1, #
        "Hours Worked Per Week": 1, #
        "With respect to the technology or technical data the petitioner will release or otherwise provide access to the beneficiary, the petitioner certifies that it has reviewed the Export Administration Regulations (EAR) and the International Traffic in Arms Regulations (ITAR) and has determined that:": 1, #Dic
        "Petitioner's or Authorized Signatory's Title": 1,
    }

    datos = []
    for archivo in os.listdir(carpeta):
        if archivo.endswith('.xlsx'):
            df = pd.read_excel(os.path.join(carpeta, archivo), header=None)
            datos_dict = {}
            for columna, fila in datos_a_obtener.items():
                if isinstance(fila, int) and fila < len(df):
                    if columna in ["startdate_proposed_employment", "enddate_proposed_employment"]:
                        # Formatear fechas
                        fecha = df.iloc[fila - 1, 1]
                        if pd.notna(fecha):
                            fecha_formateada = fecha.strftime('%m/%d/%Y')
                            datos_dict[columna] = fecha_formateada
                        else:
                            datos_dict[columna] = None
                    elif columna in ["Given Name (First Name)", "Middle Name"]:
                        # Tratar nombres
                        nombres = df.iloc[fila - 1, 1].split() if pd.notna(df.iloc[fila - 1, 1]) else [None]
                        datos_dict["Given Name (First Name)"] = nombres[0] if len(nombres) > 0 else None
                        datos_dict["Middle Name"] = nombres[1] if len(nombres) > 1 else None
                    else:
                        datos_dict[columna] = df.iloc[fila - 1, 1] if pd.notna(df.iloc[fila - 1, 1]) else None
            datos.append(datos_dict)

    df_final = pd.DataFrame(datos)
    nuevo_dataframe = fill_data(df_final)

    # Generar un nombre de archivo basado en la fecha y hora actual
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = ruta_out + f'INPUT_USERS_DATA_FORM_I-129S_{timestamp}.xlsx'

    # Crear un archivo Excel con ambas hojas
    with pd.ExcelWriter(output_path) as writer:
        nuevo_dataframe.to_excel(writer, sheet_name='Temp', index=False)

    # Color cells in the 'Temp' sheet
    wb = load_workbook(output_path)
    ws = wb['Temp']

    adjust_column_width(ws)
    apply_header_style(ws)
    apply_data_style(ws)
    
    empty_columns_yellow = [
        "startdate_proposed_employment",
        "enddate_proposed_employment",
        "Was the beneficiary of this petition in the United States during the last seven years?",
        "Country of Citizenship or Nationality",
        "Beneficiary's Wages Per Year",
        "Indicate the type of qualifying position the beneficiary was employed in while working for the qualifying foreign employer",
        "Wages Earned Per Year",
        "Hours Worked Per Week"
    ]

    empty_columns_red = [col for col in nuevo_dataframe.columns if col not in empty_columns_yellow]

    color_cells(wb, 'Temp', empty_columns_yellow, empty_columns_red)
    
    # Añadir el menú desplegable en las columnas
    add_dropdown(ws, 'N', ['Specialized Knowledge Professional', 'Manager or Executive'])
    add_dropdown(ws, 'Q', ['YES', 'NO'])
    add_dropdown(ws, 'Q', ['Manager', 'Executive', 'Specialized Knowledge Professional'])
    
    wb.save(output_path)
    
    # Mover archivos procesados a la carpeta /Casos/Procesados
    carpeta_procesados = ruta_user+'Casos/Procesados'
    for archivo in os.listdir(carpeta):
        if archivo.endswith('.xlsx'):
            shutil.move(os.path.join(carpeta, archivo), os.path.join(carpeta_procesados, archivo))

def process_files():
    global bot_running
    while bot_running:
        if any(file.endswith('.xlsx') for file in os.listdir(ruta_user+'Casos/')):
            generateExcel()
        time.sleep(10)  # Esperar 1 minuto antes de volver a comprobar

def start_bot():
    global bot_running, bot_thread
    if not bot_running:
        bot_running = True
        bot_thread = threading.Thread(target=process_files)
        bot_thread.start()

def stop_bot():
    global bot_running, bot_thread
    if bot_running:
        bot_running = False
        if bot_thread:
            bot_thread.join()
            bot_thread = None
            
            
#start_bot()