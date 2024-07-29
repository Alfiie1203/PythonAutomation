import os # Importa el módulo os para trabajar con funciones del sistema operativo
from reportlab.lib.pagesizes import letter # Importa el tamaño de página letter de reportlab
from reportlab.lib import utils # Importa utilidades de reportlab
from reportlab.pdfgen import canvas # Importa canvas para generar PDFs
import pandas as pd # Importa pandas para manipulación de datos
import openpyxl # Importa openpyxl para trabajar con archivos Excel
import threading # Importa threading para trabajar con hilos
import time # Importa time para funciones de tiempo
from openpyxl.styles import PatternFill # Importa PatternFill para estilos en celdas de Excel
import win32com.client as win32 # Importa win32com para automatización de aplicaciones COM (como Excel)
from Bot_i129s import generate_ExcelKey # Importa la función generate_ExcelKey del módulo Bot_i129s


def adjust_text_and_font_size(text, threshold_small=18, threshold_large=50, threshold_medium = 24,
                              small_font_size=9, medium_font_size=11, super_large_font_size=6.8, normaly_font_size=12):
    """
    Ajusta el texto y el tamaño de la fuente basado en la longitud del texto.
    """
    if len(text) > threshold_large:
        midpoint = (len(text) // 2) + ((len(text) // 2) // 2)
        text = text[:midpoint] + '\n' + text[midpoint:]
        return text, super_large_font_size, 5
    elif len(text) > threshold_small and len(text) < threshold_medium:
        return text, medium_font_size, 1
    elif len(text) > threshold_medium:
        return text, small_font_size, 1
    else:
        return text, normaly_font_size, 0

    
def add_text_to_image(canvas_obj, image_path, text_data):
    """
    Agrega texto a una imagen dentro de un PDF.
    """
    # Cargar la imagen
    img = utils.ImageReader(image_path)
    img_width, img_height = img.getSize()

    # Escalar la imagen para que quepa en la página
    scaling_factor = min(letter[0] / img_width, letter[1] / img_height)
    img_width *= scaling_factor
    img_height *= scaling_factor

    # Calcular la posición para centrar la imagen en la página
    x_pos = (letter[0] - img_width) / 2
    y_pos = (letter[1] - img_height) / 2

    # Agregar la imagen al lienzo
    canvas_obj.drawImage(image_path, x_pos, y_pos, width=img_width, height=img_height)

    for text_entry in text_data:
        text = text_entry['text']
        text_x = text_entry['x']
        text_y = text_entry['y']            
        
        # Ajustar el tamaño del texto y manejar saltos de línea dinámicamente
        if text_entry.get('is_add_spaces'):
            font_size = 12
            y_adjustment = 0
        else:
            text, font_size, y_adjustment = adjust_text_and_font_size(text)
        
        text_y += y_adjustment
        
        canvas_obj.setFont("Courier", font_size)
        
        # Agregar texto al lienzo en la posición especificada, manejando saltos de línea
        for line in text.split('\n'):
            canvas_obj.drawString(text_x, text_y, line)
            text_y -= font_size + 2  # Ajustar la posición y el espaciado entre líneas
            
def color_cells(workbook, sheet_name):
    """
    Colorea de rojo las celdas vacías en una hoja de Excel.
    """
    ws = workbook[sheet_name]
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        for cell in col:
            if cell.row > 1 and not cell.value:  # Verifica si la celda está vacía
                cell.fill = red_fill
 
def addSpaces(varSpaces):
    """
    Convierte el código postal en una cadena con espacios entre los dígitos.
    """
    return ' '.join(varSpaces)

def convert_date_format(date_str):
    """
    Convierte el formato de fecha de 'YYYY-DD-MM HH:MM:SS' a 'DD/MM/YYYY'.
    """
    if pd.isna(date_str) or date_str == '':
        return ''
    else: 
        date_obj = pd.to_datetime(date_str)
        return date_obj.strftime('%d/%m/%Y')
